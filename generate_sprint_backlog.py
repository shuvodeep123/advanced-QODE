"""
Generate Sprint 2-4 backlog in Excel format.

Based on the advanced-QODE repository development history and feature implementation.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Sprint 2 - Core Diagram Generation & Validation
sprint2_data = [
    # AQ-301: Process Network Diagram Validation
    ("2", "AQ-301", "AQ-301-T1", "Validate predecessor parsing and INIT-node handling", 3, "None", "Stable predecessor/INIT interpretation rules"),
    ("2", "AQ-301", "AQ-301-T2", "Verify sequential and parallel lead-time calculations", 2, "AQ-301-T1", "Correct lead-time calculation logic"),
    ("2", "AQ-301", "AQ-301-T3", "Validate critical-path logic for material-output endpoints", 2, "AQ-301-T2", "Accurate critical-path selection"),
    ("2", "AQ-301", "AQ-301-T4", "Handle duplicate, missing, and malformed activity references", 2, "AQ-301-T1", "Robust input-validation behavior"),
    ("2", "AQ-301", "AQ-301-T5", "Review node and edge labels plus DOT output quality", 2, "AQ-301-T2, AQ-301-T3", "Cleaner process diagram output"),
    
    # AQ-302: People Diagram Validation
    ("2", "AQ-302", "AQ-302-T1", "Validate role extraction from questionnaire data", 2, "None", "Reliable role extraction logic"),
    ("2", "AQ-302", "AQ-302-T2", "Verify owner-to-owner handoff mapping logic", 2, "AQ-302-T1", "Accurate people handoff mapping"),
    ("2", "AQ-302", "AQ-302-T3", "Validate automation/manual coloring and labeling rules", 2, "AQ-302-T2", "Correct visual semantics in people diagram"),
    ("2", "AQ-302", "AQ-302-T4", "Handle duplicate roles, empty owners, and self-loop cases", 2, "AQ-302-T1", "Hardened people-diagram edge cases"),
    
    # AQ-303: Technology Diagram Validation
    ("2", "AQ-303", "AQ-303-T1", "Validate automation tool extraction and normalization", 2, "None", "Clean tool inventory for diagraming"),
    ("2", "AQ-303", "AQ-303-T2", "Verify same-tool and cross-tool dependency mapping", 2, "AQ-303-T1", "Accurate technology flow mapping"),
    ("2", "AQ-303", "AQ-303-T3", "Improve edge deduplication and aggregated edge labeling", 2, "AQ-303-T2", "Reduced duplicate edges in technology diagram"),
    ("2", "AQ-303", "AQ-303-T4", "Handle missing tools, aliases, and invalid tool-pair relationships", 2, "AQ-303-T1", "More resilient technology diagram generation"),
    
    # AQ-304: Multi-format Diagram Export
    ("2", "AQ-304", "AQ-304-T1", "Align DOT, Mermaid, and PlantUML generation behavior", 2, "AQ-301, AQ-302, AQ-303", "Consistent multi-format diagram outputs"),
    ("2", "AQ-304", "AQ-304-T2", "Standardize PNG rendering and export output checks", 2, "AQ-304-T1", "Reliable rendered diagram artifacts"),
    ("2", "AQ-304", "AQ-304-T3", "Validate download naming and artifact directory structure", 1, "AQ-304-T2", "Standardized artifact naming/location"),
    ("2", "AQ-304", "AQ-304-T4", "Confirm reviewer-friendly consistency across architecture types", 1, "AQ-304-T1", "Consistent user-facing output set"),
    
    # AQ-305: Document Export Capability
    ("2", "AQ-305", "AQ-305-T1", "Standardize titles, headers, and metadata in exports", 2, "AQ-304-T1", "Uniform export metadata"),
    ("2", "AQ-305", "AQ-305-T2", "Improve formatting for lists, sections, and grouping", 1, "AQ-305-T1", "Better document readability"),
    ("2", "AQ-305", "AQ-305-T3", "Validate narrative alignment with generated architecture type", 1, "AQ-305-T1", "Correct narrative-to-diagram alignment"),
    ("2", "AQ-305", "AQ-305-T4", "Verify file naming, versioning, and MIME/download behavior", 1, "AQ-305-T1", "Download-ready export package"),
]

# Sprint 3 - RAG Pipeline Foundation
sprint3_data = [
    # AQ-201: Document Ingestion Pipeline
    ("3", "AQ-201", "AQ-201-T1", "Validate required questionnaire columns and sheet assumptions", 2, "None", "Ingestion input contract"),
    ("3", "AQ-201", "AQ-201-T2", "Improve malformed and incomplete Excel failure handling", 2, "AQ-201-T1", "Better ingestion error handling"),
    ("3", "AQ-201", "AQ-201-T3", "Strengthen parsing for xlsm, xlsx, docx, pdf, and txt inputs", 2, "AQ-201-T1", "Multi-format ingestion support"),
    ("3", "AQ-201", "AQ-201-T4", "Normalize extracted text and metadata before chunking", 2, "AQ-201-T3", "Clean normalized ingestion payloads"),
    ("3", "AQ-201", "AQ-201-T5", "Ensure bad files are skipped without breaking the session", 1, "AQ-201-T2, AQ-201-T3", "Session-safe ingestion behavior"),
    
    # AQ-202: ChromaDB Vector Store Integration
    ("3", "AQ-202", "AQ-202-T1", "Review chunk size and overlap strategy across sources", 2, "AQ-201-T4", "Improved chunking configuration"),
    ("3", "AQ-202", "AQ-202-T2", "Reduce redundant and low-value chunks", 2, "AQ-202-T1", "Higher-value retrieval corpus"),
    ("3", "AQ-202", "AQ-202-T3", "Improve chunk boundary quality for coherent retrieval", 1, "AQ-202-T1", "Better semantic chunk integrity"),
    ("3", "AQ-202", "AQ-202-T4", "Preserve source attribution and traceability in chunks", 1, "AQ-202-T2", "Source-aware retrieval metadata"),
    
    # AQ-203: Knowledge Graph Construction
    ("3", "AQ-203", "AQ-203-T1", "Validate node creation for pillars, roles, tools, and activities", 2, "AQ-201-T4", "Complete knowledge graph node model"),
    ("3", "AQ-203", "AQ-203-T2", "Verify relationship mapping across core graph edges", 2, "AQ-203-T1", "Correct graph relationships"),
    ("3", "AQ-203", "AQ-203-T3", "Handle partial or missing questionnaire data during graph build", 2, "AQ-203-T1", "Resilient graph build process"),
    ("3", "AQ-203", "AQ-203-T4", "Validate graph persistence, reload, and integrity checks", 2, "AQ-203-T2", "Persisted and reloadable graph store"),
    
    # AQ-204: Hybrid Retrieval Implementation
    ("3", "AQ-204", "AQ-204-T1", "Refine intent-based retrieval filtering by discipline", 2, "AQ-202, AQ-203", "Better query-to-context routing"),
    ("3", "AQ-204", "AQ-204-T2", "Improve ranking for People, Process, and Technology queries", 2, "AQ-204-T1", "Better relevance ordering"),
    ("3", "AQ-204", "AQ-204-T3", "Tune graph and vector retrieval balance to reduce noise", 2, "AQ-203-T4, AQ-204-T2", "Hybrid retrieval optimization"),
    ("3", "AQ-204", "AQ-204-T4", "Validate evidence alignment with question and architecture type", 1, "AQ-204-T3", "Grounded response quality checks"),
    
    # AQ-205: RAG Evaluation Framework
    ("3", "AQ-205", "AQ-205-T1", "Define score-threshold rules for answer quality", 2, "AQ-204-T4", "Response quality baseline"),
    ("3", "AQ-205", "AQ-205-T2", "Create regression scenarios for common prompt types", 1, "AQ-205-T1", "Reusable evaluation scenarios"),
    ("3", "AQ-205", "AQ-205-T3", "Compare grounded versus weak-context response quality", 1, "AQ-205-T2", "Comparative retrieval-quality evidence"),
    ("3", "AQ-205", "AQ-205-T4", "Publish repeatable evaluation checks for future sprints", 1, "AQ-205-T1, AQ-205-T2", "Repeatable validation checklist"),
]

# Sprint 4 - Advanced Features & Production Readiness
sprint4_data = [
    # AQ-401: Engineering Principles Framework
    ("4", "AQ-401", "AQ-401-T1", "Define the 9 engineering principles to 3 disciplines mapping model", 2, "AQ-203", "Principle-to-discipline mapping baseline"),
    ("4", "AQ-401", "AQ-401-T2", "Create targeted context enrichment for People, Process, and Technology", 2, "AQ-401-T1", "Principle-aware prompt enrichment"),
    ("4", "AQ-401", "AQ-401-T3", "Add principle extraction and request classification for To-Be analysis", 2, "AQ-401-T2", "To-Be routing logic"),
    ("4", "AQ-401", "AQ-401-T4", "Validate principle coverage for common transformation prompts", 1, "AQ-401-T3", "Coverage validation results"),
    
    # AQ-402: Graph-RAG Enhancements
    ("4", "AQ-402", "AQ-402-T1", "Add entity extraction for pillars, roles, and tools from user queries", 2, "AQ-203-T4", "Query entity extraction layer"),
    ("4", "AQ-402", "AQ-402-T2", "Build multi-hop graph traversal for structural context retrieval", 2, "AQ-402-T1", "Graph traversal retrieval"),
    ("4", "AQ-402", "AQ-402-T3", "Merge graph and vector context with deduplication and source tagging", 2, "AQ-402-T2, AQ-204-T3", "Hybrid Graph-RAG retriever"),
    ("4", "AQ-402", "AQ-402-T4", "Add global-summary fallback when no direct entity match exists", 1, "AQ-402-T3", "Fallback retrieval behavior"),
    
    # AQ-403: To-Be & Gap Analysis Diagrams
    ("4", "AQ-403", "AQ-403-T1", "Generate To-Be diagrams in DOT and Mermaid formats from LLM output", 3, "AQ-401, AQ-402", "To-Be diagram generation pipeline"),
    ("4", "AQ-403", "AQ-403-T2", "Derive PlantUML artifacts from generated architecture outputs", 2, "AQ-403-T1", "PlantUML diagram export"),
    ("4", "AQ-403", "AQ-403-T3", "Support gap-analysis mode for removed pillars or components", 2, "AQ-403-T1", "Gap-analysis flow"),
    ("4", "AQ-403", "AQ-403-T4", "Highlight direct and indirect impact nodes in gap diagrams", 2, "AQ-403-T3", "Risk-highlighted gap diagrams"),
    
    # AQ-404: Multi-Format Document Export
    ("4", "AQ-404", "AQ-404-T1", "Convert long-form LLM analysis into structured document sections", 2, "AQ-401-T3", "Structured narrative content blocks"),
    ("4", "AQ-404", "AQ-404-T2", "Enable downloadable Word and Excel exports for analysis output", 2, "AQ-404-T1", "Word and Excel export support"),
    ("4", "AQ-404", "AQ-404-T3", "Enable downloadable PowerPoint and PDF exports for analysis output", 2, "AQ-404-T1", "PowerPoint and PDF export support"),
    ("4", "AQ-404", "AQ-404-T4", "Validate file naming, MIME types, and preview/download behavior", 1, "AQ-404-T2, AQ-404-T3", "Production-ready export downloads"),
    
    # AQ-405: Production Features
    ("4", "AQ-405", "AQ-405-T1", "Add persistent chat history for multi-session continuity", 2, "AQ-402", "Session history capability"),
    ("4", "AQ-405", "AQ-405-T2", "Add token usage tracking and budget visibility in the UI", 2, "AQ-405-T1", "Token usage dashboard"),
    ("4", "AQ-405", "AQ-405-T3", "Add Langfuse-based observability and trace capture for LLM flows", 2, "AQ-402, AQ-405-T2", "LLM traceability"),
    ("4", "AQ-405", "AQ-405-T4", "Refine UI experience for upload, mode badges, and download actions", 1, "AQ-404-T4, AQ-405-T1", "Improved user interaction flow"),
]

# Combine all sprint data
all_data = sprint2_data + sprint3_data + sprint4_data

# Create DataFrame
columns = ["Sprint No.", "Story ID", "Task ID", "Task Description", "Task Story Points", "Dependencies", "Deliverables"]
df = pd.DataFrame(all_data, columns=columns)

# Create Excel workbook with formatting
wb = Workbook()
ws = wb.active
ws.title = "Sprint 2-4 Backlog"

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

story_fill_s2 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
story_fill_s3 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
story_fill_s4 = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

cell_alignment = Alignment(vertical="top", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
column_widths = [12, 12, 12, 50, 15, 20, 40]
for idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[chr(64 + idx)].width = width

# Add header row
for col_idx, column_name in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=column_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add data rows
for row_idx, row_data in enumerate(all_data, 2):
    sprint_no = row_data[0]
    
    # Determine fill color based on sprint
    if sprint_no == "2":
        row_fill = story_fill_s2
    elif sprint_no == "3":
        row_fill = story_fill_s3
    else:
        row_fill = story_fill_s4
    
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = cell_alignment
        cell.border = border
        cell.fill = row_fill

# Freeze header row
ws.freeze_panes = "A2"

# Add filters
ws.auto_filter.ref = ws.dimensions

# Save workbook
output_file = "/home/runner/work/advanced-QODE/advanced-QODE/Sprint_2-4_Backlog.xlsx"
wb.save(output_file)

print(f"✓ Sprint 2-4 backlog exported to: {output_file}")
print(f"✓ Total stories: {len(set(row[1] for row in all_data))}")
print(f"✓ Total tasks: {len(all_data)}")
print(f"✓ Sprint 2 tasks: {len(sprint2_data)}")
print(f"✓ Sprint 3 tasks: {len(sprint3_data)}")
print(f"✓ Sprint 4 tasks: {len(sprint4_data)}")
