"""
diagram_executor.py — Bridge between the RAG chain and the QODE diagram generators.

Output formats supported
------------------------
  1. PNG   — rendered via pydot / Graphviz (preferred when Graphviz is installed).
  2. Mermaid (.mmd)  — converted from DOT graph using the dot→flowchart translator;
                       renderable in-browser via mermaid.js (no install needed).
  3. PlantUML (.puml) — converted from DOT graph; renderable via PlantUML server.

The executor tries formats in priority order: PNG → Mermaid → PlantUML → raw DOT.
The first successful output path is returned.

Public API
----------
    run(diagram_type, excel_path, output_format) -> str | None
        Returns absolute path to the generated file, or None on failure.

    dot_to_mermaid(dot_text, diagram_type) -> str
        Convert a Graphviz DOT string to Mermaid flowchart syntax.

    dot_to_plantuml(dot_text, diagram_type) -> str
        Convert a Graphviz DOT string to PlantUML activity/component syntax.
"""

from __future__ import annotations

import logging
import os
import re
import sys
from pathlib import Path

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Diagram type → (module, class, method, dot_output_stem)
# ---------------------------------------------------------------------------
_DIAGRAM_MAP: dict[str, tuple[str, str, str, str]] = {
    "process": (
        "Generate_Process_Network_Diagram",
        "Process_Diagram",
        "create_network_diagram",
        "Diagram_Network",
    ),
    "people": (
        "Generate_People_Diagram",
        "People_Diagram",
        "create_people_diagram",
        "Diagram_People",
    ),
    "technology": (
        "Generate_Technology_Diagram",
        "Technology_Diagram",
        "create_technology_diagram",
        "Diagram_Technology",
    ),
}

_REPO_ROOT = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------------
# Structured output directories
# AS-IS and TO-BE diagrams are persisted in their own sub-trees so they are
# easy to find, version, and share independently.
# ---------------------------------------------------------------------------
ASIS_DIRS: dict[str, Path] = {
    "dot":      _REPO_ROOT / "AS-IS" / "DotGraph",
    "mermaid":  _REPO_ROOT / "AS-IS" / "Mermaid",
    "drawio":   _REPO_ROOT / "AS-IS" / "draw.io",
    "plantuml": _REPO_ROOT / "AS-IS" / "PlantUML",
}
TOBE_DIRS: dict[str, Path] = {
    "dot":      _REPO_ROOT / "TO-BE" / "DotGraph",
    "mermaid":  _REPO_ROOT / "TO-BE" / "Mermaid",
    "drawio":   _REPO_ROOT / "TO-BE" / "draw.io",
    "plantuml": _REPO_ROOT / "TO-BE" / "PlantUML",
}


def _ensure_output_dirs() -> None:
    """Create all AS-IS and TO-BE output directories if they don't exist."""
    for d in list(ASIS_DIRS.values()) + list(TOBE_DIRS.values()):
        d.mkdir(parents=True, exist_ok=True)


_ensure_output_dirs()

# ---------------------------------------------------------------------------
# QODE questionnaire discovery — accepts any .xlsm, not just a hardcoded name
# ---------------------------------------------------------------------------

# Sheet-name hints that indicate a workbook is QODE-compatible
_QODE_SHEET_HINTS: frozenset[str] = frozenset(
    {
        "people", "process", "technology", "q_stories", "qode",
        "questionnaire", "pillars", "roles", "tools", "activities",
    }
)


def _score_xlsm(p: Path) -> int:
    """Return the number of QODE sheet-name hints found in *p*. 0 if unreadable."""
    try:
        import openpyxl  # type: ignore[import]
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        sheet_names_lower = {s.lower() for s in wb.sheetnames}
        wb.close()
        return sum(1 for hint in _QODE_SHEET_HINTS if hint in sheet_names_lower)
    except Exception:
        return 0


def find_questionnaire(excel_path: str | None = None) -> str | None:
    """Locate a QODE-compatible Excel workbook.

    Priority
    --------
    1. The *excel_path* argument — if provided **and** the file exists.
    2. ``QODE-Questionnaire.xlsm`` in the repo root (legacy default name).
    3. Any ``.xlsm / .xlsx`` in the repo root ranked by QODE-sheet coverage.
    4. First ``.xlsm / .xlsx`` found anywhere in the repo root (last resort).

    Returns the resolved absolute path string, or ``None`` when no file is found.
    """
    # 1. Explicit path
    if excel_path:
        p = Path(excel_path).resolve()
        if p.exists() and p.suffix.lower() in (".xlsm", ".xlsx"):
            logger.info("Using explicit questionnaire: %s", p)
            return str(p)
        logger.warning(
            "Explicit excel_path '%s' not found or not xlsm/xlsx — scanning repo root.",
            excel_path,
        )

    # 2. Legacy default name
    legacy = _REPO_ROOT / "QODE-Questionnaire.xlsm"
    if legacy.exists():
        logger.info("Using legacy questionnaire: %s", legacy)
        return str(legacy)

    # 3 & 4. Scan repo root
    candidates = list(_REPO_ROOT.glob("*.xlsm")) + list(_REPO_ROOT.glob("*.xlsx"))
    if not candidates:
        logger.warning("No .xlsm/.xlsx files found in repo root '%s'.", _REPO_ROOT)
        return None
    if len(candidates) == 1:
        logger.info("Using only available questionnaire: %s", candidates[0])
        return str(candidates[0].resolve())

    best = max(candidates, key=_score_xlsm)
    logger.info("Selected questionnaire by QODE-sheet score: %s", best)
    return str(best.resolve())


# ---------------------------------------------------------------------------
# Path helper — used by chain.py to read As-Is DOT after generation
# ---------------------------------------------------------------------------

def get_dot_path(diagram_type: str) -> Path | None:
    """Return the expected DOT file path for *diagram_type*, or None if unknown."""
    entry = _DIAGRAM_MAP.get(diagram_type)
    if not entry:
        return None
    return _REPO_ROOT / entry[3]  # dot_filename stem (no extension)


# ---------------------------------------------------------------------------
# DOT → Mermaid converter
# ---------------------------------------------------------------------------

def dot_to_mermaid(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to Mermaid flowchart syntax.

    Handles:
      - Directed (digraph) and undirected (graph) graphs
      - Node labels (quoted strings)
      - Edge labels
      - Node fill-colour → Mermaid style classes
      - Basic subgraph blocks (flattened)

    Guarantees: Returns syntactically valid Mermaid with escaped special chars.
    """
    lines: list[str] = ["flowchart TD"]

    # Strip comments
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)

    # Determine direction
    if "rankdir=LR" in dot or "rankdir = LR" in dot:
        lines[0] = "flowchart LR"

    node_ids: dict[str, str] = {}   # raw_id → safe_id
    node_labels: dict[str, str] = {}
    node_colors: dict[str, str] = {}
    edge_count = 0

    def _safe_id(raw: str) -> str:
        """Convert a DOT node id to a Mermaid-safe alphanumeric id."""
        s = re.sub(r"[^a-zA-Z0-9_]", "_", raw.strip('"'))
        if s and s[0].isdigit():
            s = "N" + s
        return s or "NODE"

    def _safe_label(lbl: str) -> str:
        """Escape label for Mermaid syntax safety."""
        lbl = lbl.strip('"').strip("'")
        # Remove/replace problematic chars for Mermaid
        lbl = lbl.replace("[", "(").replace("]", ")").replace("{", "(").replace("}", ")")
        lbl = lbl.replace('"', "'")  # replace double quotes with single
        # Truncate very long labels
        if len(lbl) > 80:
            lbl = lbl[:77] + "..."
        return lbl

    is_directed = "digraph" in dot.lower()
    arrow = "-->" if is_directed else "---"

    # ── Parse node statements ──────────────────────────────────────────────
    # Pattern: id [label="...", fillcolor="...", ...]
    node_pattern = re.compile(
        r'^\s*"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE
    )
    for m in node_pattern.finditer(dot):
        raw_id = m.group(1)
        attrs_str = m.group(2)
        safe = _safe_id(raw_id)
        node_ids[raw_id] = safe

        # Extract label
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs_str)
        label = _safe_label(lbl_m.group(1)) if lbl_m else _safe_label(raw_id)
        node_labels[raw_id] = label

        # Extract fillcolor for style
        col_m = re.search(r'fillcolor\s*=\s*"?([^",\]]+)"?', attrs_str)
        if col_m:
            node_colors[raw_id] = col_m.group(1).strip()

    # ── Parse edge statements ──────────────────────────────────────────────
    # Pattern: id1 -> id2 [label="..."]  or  id1 -- id2
    edge_sep = r"->" if is_directed else r"--"
    edge_pattern = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pattern.finditer(dot):
        src_raw = m.group(1).strip()
        dst_raw = m.group(2).strip()
        attrs_str = m.group(3) or ""

        src = node_ids.get(src_raw, _safe_id(src_raw))
        dst = node_ids.get(dst_raw, _safe_id(dst_raw))

        # Ensure nodes are declared
        if src_raw not in node_labels:
            node_labels[src_raw] = _safe_label(src_raw)
            node_ids[src_raw] = src
        if dst_raw not in node_labels:
            node_labels[dst_raw] = _safe_label(dst_raw)
            node_ids[dst_raw] = dst

        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs_str)
        edge_label = _safe_label(lbl_m.group(1)) if lbl_m else ""

        src_label = node_labels.get(src_raw, src)
        dst_label = node_labels.get(dst_raw, dst)

        if edge_label:
            lines.append(f'    {src}["{src_label}"] {arrow}|"{edge_label}"| {dst}["{dst_label}"]')
        else:
            lines.append(f'    {src}["{src_label}"] {arrow} {dst}["{dst_label}"]')
        edge_count += 1

    # ── Emit isolated nodes (no edges) ────────────────────────────────────
    for raw_id, safe in node_ids.items():
        label = node_labels.get(raw_id, raw_id)
        # Only add if not already in an edge line
        node_ref = f'{safe}['
        if not any(node_ref in l for l in lines[1:]):
            lines.append(f'    {safe}["{label}"]')

    # ── Style classes from fillcolors ──────────────────────────────────────
    color_classes: dict[str, list[str]] = {}
    for raw_id, color in node_colors.items():
        safe = node_ids.get(raw_id, _safe_id(raw_id))
        color_classes.setdefault(color, []).append(safe)

    for color, ids in color_classes.items():
        cls_name = "cls_" + re.sub(r"[^a-zA-Z0-9]", "_", color)
        lines.append(f'    classDef {cls_name} fill:{color},stroke:#333,color:#fff')
        lines.append(f'    class {",".join(ids)} {cls_name}')

    if edge_count == 0 and not node_ids:
        # Fallback: no parseable structure
        lines.append('    A["Diagram data not parsed"]')

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# DOT → PlantUML converter
# ---------------------------------------------------------------------------

def dot_to_plantuml(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to PlantUML component/activity syntax."""
    is_directed = "digraph" in dot_text.lower()

    # Strip comments
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)

    puml_type = "component" if diagram_type == "technology" else "activity" if diagram_type == "process" else "usecase"

    lines = ["@startuml", f"' {diagram_type.title()} Architecture Diagram", ""]

    if diagram_type == "people":
        lines += ["left to right direction", ""]

    node_labels: dict[str, str] = {}
    edges: list[tuple[str, str, str]] = []

    def _label(raw: str) -> str:
        return raw.strip('"').strip("'").replace("<", "").replace(">", "")

    # Parse nodes
    node_pat = re.compile(r'"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE)
    for m in node_pat.finditer(dot):
        raw_id = m.group(1)
        attrs = m.group(2)
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        label = _label(lbl_m.group(1)) if lbl_m else _label(raw_id)
        node_labels[raw_id] = label

    # Parse edges
    edge_sep = "->" if is_directed else "--"
    edge_pat = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pat.finditer(dot):
        src = m.group(1).strip()
        dst = m.group(2).strip()
        attrs = m.group(3) or ""
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        elabel = _label(lbl_m.group(1)) if lbl_m else ""
        edges.append((src, dst, elabel))
        if src not in node_labels:
            node_labels[src] = _label(src)
        if dst not in node_labels:
            node_labels[dst] = _label(dst)

    arrow = "-->" if is_directed else "--"

    for raw_id, label in node_labels.items():
        safe_id = re.sub(r"[^a-zA-Z0-9]", "_", raw_id)
        if diagram_type == "technology":
            lines.append(f'component "{label}" as {safe_id}')
        elif diagram_type == "people":
            lines.append(f'actor "{label}" as {safe_id}')
        else:  # process — use rectangle so the full network structure is preserved
            lines.append(f'rectangle "{label}" as {safe_id}')

    lines.append("")

    for src, dst, elabel in edges:
        s = re.sub(r"[^a-zA-Z0-9]", "_", src)
        d = re.sub(r"[^a-zA-Z0-9]", "_", dst)
        lbl_part = f" : {elabel}" if elabel else ""
        lines.append(f"{s} {arrow} {d}{lbl_part}")

    lines.append("")
    lines.append("@enduml")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# DOT → draw.io XML converter
# ---------------------------------------------------------------------------

def dot_to_drawio(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to draw.io (diagrams.net) XML.

    The returned XML string can be imported directly into draw.io / diagrams.net
    via File → Import from → XML, or saved as a ``.drawio`` file and opened.
    """
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)
    is_directed = "digraph" in dot.lower()

    def _lbl(raw: str) -> str:
        return raw.strip('"').strip("'")

    node_labels: dict[str, str] = {}
    node_colors: dict[str, str] = {}
    edges: list[tuple[str, str, str]] = []

    node_pat = re.compile(r'"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE)
    for m in node_pat.finditer(dot):
        raw_id, attrs = m.group(1), m.group(2)
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        node_labels[raw_id] = _lbl(lbl_m.group(1)) if lbl_m else _lbl(raw_id)
        col_m = re.search(r'fillcolor\s*=\s*"?([^",\]]+)"?', attrs)
        if col_m:
            node_colors[raw_id] = col_m.group(1).strip()

    edge_sep = "->" if is_directed else "--"
    edge_pat = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pat.finditer(dot):
        src, dst = m.group(1).strip(), m.group(2).strip()
        attrs = m.group(3) or ""
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        edges.append((src, dst, _lbl(lbl_m.group(1)) if lbl_m else ""))
        node_labels.setdefault(src, _lbl(src))
        node_labels.setdefault(dst, _lbl(dst))

    # Style per diagram type
    _STYLE = {
        "process":    "rounded=1;whiteSpace=wrap;html=1;fillColor=#dae8fc;strokeColor=#6c8ebf;",
        "people":     "shape=mxgraph.basic.person;whiteSpace=wrap;html=1;fillColor=#f8cecc;strokeColor=#b85450;",
        "technology": "rounded=1;arcSize=20;whiteSpace=wrap;html=1;fillColor=#d5e8d4;strokeColor=#82b366;",
    }
    base_style = _STYLE.get(diagram_type, _STYLE["process"])
    edge_style = (
        "rounded=0;orthogonalLoop=1;jettySize=auto;exitX=0.5;exitY=1;entryX=0.5;entryY=0;"
        + ("endArrow=block;endFill=1;" if is_directed else "endArrow=none;")
    )

    # Layout: left-right for technology, top-down otherwise
    cols = max(1, int(len(node_labels) ** 0.5) + 1)
    x_gap, y_gap = 160, 100
    x_start, y_start = 80, 80

    cells: list[str] = [
        '<mxCell id="0"/>',
        '<mxCell id="1" parent="0"/>',
    ]
    node_ids: dict[str, str] = {}
    cell_id = 2

    for i, (raw_id, label) in enumerate(node_labels.items()):
        col, row = i % cols, i // cols
        x, y = x_start + col * x_gap, y_start + row * y_gap
        color_part = (
            f'fillColor={node_colors[raw_id]};'
            if raw_id in node_colors else ""
        )
        cid = str(cell_id)
        node_ids[raw_id] = cid
        safe_label = label.replace('"', '&quot;').replace('<', '&lt;').replace('>', '&gt;')
        cells.append(
            f'<mxCell id="{cid}" value="{safe_label}" style="{base_style}{color_part}" '
            f'vertex="1" parent="1"><mxGeometry x="{x}" y="{y}" width="120" height="60" '
            f'as="geometry"/></mxCell>'
        )
        cell_id += 1

    for src, dst, elabel in edges:
        src_id = node_ids.get(src, "1")
        dst_id = node_ids.get(dst, "1")
        safe_elabel = elabel.replace('"', '&quot;')
        cells.append(
            f'<mxCell id="{cell_id}" value="{safe_elabel}" style="{edge_style}" '
            f'edge="1" source="{src_id}" target="{dst_id}" parent="1">'
            f'<mxGeometry relative="1" as="geometry"/></mxCell>'
        )
        cell_id += 1

    cells_xml = "\n        ".join(cells)
    title = f"{diagram_type.title()} Architecture"
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<mxfile version="20.0.0">\n'
        f'  <diagram name="{title}">\n'
        f'    <mxGraphModel dx="1034" dy="546" grid="1" gridSize="10" guides="1" '
        f'tooltips="1" connect="1" arrows="1" fold="1" page="0" pageScale="1" '
        f'pageWidth="1169" pageHeight="827" math="0" shadow="0">\n'
        f'      <root>\n'
        f'        {cells_xml}\n'
        f'      </root>\n'
        f'    </mxGraphModel>\n'
        f'  </diagram>\n'
        f'</mxfile>'
    )


# ---------------------------------------------------------------------------
# Mermaid syntax validation
# ---------------------------------------------------------------------------

def _validate_mermaid(mmd_text: str) -> tuple[bool, str]:
    """Validate Mermaid syntax. Returns (is_valid, error_msg)."""
    if not mmd_text or not isinstance(mmd_text, str):
        return False, "Empty or invalid diagram text"

    lines = [l.strip() for l in mmd_text.split('\n') if l.strip()]
    if not lines:
        return False, "Empty diagram"

    # Check first line is a valid diagram declaration
    first = lines[0].lower()
    valid_types = ('flowchart', 'graph', 'sequencediagram', 'classdiagram', 'statediagram')
    if not any(first.startswith(d) for d in valid_types):
        return False, f"Invalid diagram type: {first}"

    # Basic bracket matching (strict: count all bracket types)
    open_square = mmd_text.count('[')
    close_square = mmd_text.count(']')
    open_paren = mmd_text.count('(')
    close_paren = mmd_text.count(')')
    open_curly = mmd_text.count('{')
    close_curly = mmd_text.count('}')

    if open_square != close_square:
        return False, f"Unbalanced square brackets: {open_square} open, {close_square} close"
    if open_paren != close_paren:
        return False, f"Unbalanced parentheses: {open_paren} open, {close_paren} close"
    if open_curly != close_curly:
        return False, f"Unbalanced braces: {open_curly} open, {close_curly} close"

    # Safety: reject if too large (>100k chars suggests infinite loop or DoS)
    if len(mmd_text) > 100000:
        return False, "Diagram exceeds size limit (100k chars)"

    return True, ""


# ---------------------------------------------------------------------------
# Main executor
# ---------------------------------------------------------------------------

def run(
    diagram_type: str,
    excel_path: str | None = None,
    output_format: str = "mermaid",   # "mermaid" only; stored formats generated as artifacts
    is_tobe: bool = False,            # AS-IS (False) or TO-BE (True)
) -> str | None:
    """Generate a QODE diagram and return the Mermaid path.

    Always returns Mermaid .mmd format. Artifacts in other formats (DOT, PlantUML, draw.io)
    are generated for archival but not returned.

    Args:
        diagram_type:   "process" | "people" | "technology"
        excel_path:     Path to the QODE questionnaire. Defaults to repo root default.
        output_format:  Ignored; always returns Mermaid.
        is_tobe:        True = TO-BE diagrams, False = AS-IS diagrams (default).

    Returns:
        Absolute path to the Mermaid .mmd file, or None on failure.
    """
    if diagram_type not in _DIAGRAM_MAP:
        logger.error("Unknown diagram_type '%s'", diagram_type)
        return None

    module_name, class_name, method_name, dot_filename = _DIAGRAM_MAP[diagram_type]

    # Smart questionnaire resolution — accepts any .xlsm, not just the hardcoded name
    resolved_excel = find_questionnaire(excel_path)
    if resolved_excel is None:
        logger.error(
            "No QODE questionnaire file found. "
            "Upload a .xlsm/.xlsx file or place one in the repo root."
        )
        return None
    excel_abs = resolved_excel

    original_cwd = os.getcwd()
    os.chdir(str(_REPO_ROOT))

    # Ensure the generator modules always find the file under the canonical name
    expected_path = _REPO_ROOT / "QODE-Questionnaire.xlsm"
    if Path(excel_abs).resolve() != expected_path.resolve() and Path(excel_abs).exists():
        import shutil
        shutil.copy(excel_abs, str(expected_path))

    try:
        repo_str = str(_REPO_ROOT)
        if repo_str not in sys.path:
            sys.path.insert(0, repo_str)

        if module_name in sys.modules:
            del sys.modules[module_name]

        import importlib
        module = importlib.import_module(module_name)
        instance = getattr(module, class_name)()
        getattr(instance, method_name)()

        dot_path = _REPO_ROOT / dot_filename
        if not dot_path.exists():
            logger.warning("DOT file '%s' was not created.", dot_path)
            return None

        dot_text = dot_path.read_text(encoding="utf-8", errors="replace")

        output_dirs = TOBE_DIRS if is_tobe else ASIS_DIRS

        # ── Save DOT to structured dir ────────────────────────────────────
        try:
            dot_out = output_dirs["dot"] / f"{dot_filename}.dot"
            dot_out.write_text(dot_text, encoding="utf-8")
        except Exception as _e:
            logger.warning("Could not write DOT to structured dir: %s", _e)

        # ── Primary path: Mermaid (.mmd) ──────────────────────────────────
        mmd_out = output_dirs["mermaid"] / f"{dot_filename}.mmd"
        try:
            mmd_text = dot_to_mermaid(dot_text, diagram_type)
            is_valid, err_msg = _validate_mermaid(mmd_text)
            if not is_valid:
                logger.error("Mermaid validation failed: %s", err_msg)
                return None
            mmd_out.write_text(mmd_text, encoding="utf-8")
            logger.info("Mermaid diagram saved: %s", mmd_out)

            # Generate artifact formats (not returned to UI)
            # ── PlantUML artifact
            try:
                puml_text = dot_to_plantuml(dot_text, diagram_type)
                puml_out = output_dirs["plantuml"] / f"{dot_filename}.puml"
                puml_out.write_text(puml_text, encoding="utf-8")
                logger.info("PlantUML artifact saved: %s", puml_out)
            except Exception:
                pass

            # ── draw.io artifact
            try:
                drawio_text = dot_to_drawio(dot_text, diagram_type)
                drawio_out = output_dirs["drawio"] / f"{dot_filename}.drawio"
                drawio_out.write_text(drawio_text, encoding="utf-8")
                logger.info("draw.io artifact saved: %s", drawio_out)
            except Exception:
                pass

            return str(mmd_out)
        except Exception as mmd_err:
            logger.error("Mermaid conversion failed: %s", mmd_err)
            return None

    except Exception as exc:
        logger.error("Diagram generation failed for '%s': %s", diagram_type, exc)
        return None

    finally:
        os.chdir(original_cwd)
# ---------------------------------------------------------------------------
ASIS_DIRS: dict[str, Path] = {
    "dot":      _REPO_ROOT / "AS-IS" / "DotGraph",
    "mermaid":  _REPO_ROOT / "AS-IS" / "Mermaid",
    "drawio":   _REPO_ROOT / "AS-IS" / "draw.io",
    "plantuml": _REPO_ROOT / "AS-IS" / "PlantUML",
}
TOBE_DIRS: dict[str, Path] = {
    "dot":      _REPO_ROOT / "TO-BE" / "DotGraph",
    "mermaid":  _REPO_ROOT / "TO-BE" / "Mermaid",
    "drawio":   _REPO_ROOT / "TO-BE" / "draw.io",
    "plantuml": _REPO_ROOT / "TO-BE" / "PlantUML",
}


def _ensure_output_dirs() -> None:
    """Create all AS-IS and TO-BE output directories if they don't exist."""
    for d in list(ASIS_DIRS.values()) + list(TOBE_DIRS.values()):
        d.mkdir(parents=True, exist_ok=True)


_ensure_output_dirs()

# ---------------------------------------------------------------------------
# QODE questionnaire discovery — accepts any .xlsm, not just a hardcoded name
# ---------------------------------------------------------------------------

# Sheet-name hints that indicate a workbook is QODE-compatible
_QODE_SHEET_HINTS: frozenset[str] = frozenset(
    {
        "people", "process", "technology", "q_stories", "qode",
        "questionnaire", "pillars", "roles", "tools", "activities",
    }
)


def _score_xlsm(p: Path) -> int:
    """Return the number of QODE sheet-name hints found in *p*. 0 if unreadable."""
    try:
        import openpyxl  # type: ignore[import]
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        sheet_names_lower = {s.lower() for s in wb.sheetnames}
        wb.close()
        return sum(1 for hint in _QODE_SHEET_HINTS if hint in sheet_names_lower)
    except Exception:
        return 0


def find_questionnaire(excel_path: str | None = None) -> str | None:
    """Locate a QODE-compatible Excel workbook.

    Priority
    --------
    1. The *excel_path* argument — if provided **and** the file exists.
    2. ``QODE-Questionnaire.xlsm`` in the repo root (legacy default name).
    3. Any ``.xlsm / .xlsx`` in the repo root ranked by QODE-sheet coverage.
    4. First ``.xlsm / .xlsx`` found anywhere in the repo root (last resort).

    Returns the resolved absolute path string, or ``None`` when no file is found.
    """
    # 1. Explicit path
    if excel_path:
        p = Path(excel_path).resolve()
        if p.exists() and p.suffix.lower() in (".xlsm", ".xlsx"):
            logger.info("Using explicit questionnaire: %s", p)
            return str(p)
        logger.warning(
            "Explicit excel_path '%s' not found or not xlsm/xlsx — scanning repo root.",
            excel_path,
        )

    # 2. Legacy default name
    legacy = _REPO_ROOT / "QODE-Questionnaire.xlsm"
    if legacy.exists():
        logger.info("Using legacy questionnaire: %s", legacy)
        return str(legacy)

    # 3 & 4. Scan repo root
    candidates = list(_REPO_ROOT.glob("*.xlsm")) + list(_REPO_ROOT.glob("*.xlsx"))
    if not candidates:
        logger.warning("No .xlsm/.xlsx files found in repo root '%s'.", _REPO_ROOT)
        return None
    if len(candidates) == 1:
        logger.info("Using only available questionnaire: %s", candidates[0])
        return str(candidates[0].resolve())

    best = max(candidates, key=_score_xlsm)
    logger.info("Selected questionnaire by QODE-sheet score: %s", best)
    return str(best.resolve())


# ---------------------------------------------------------------------------
# Path helper — used by chain.py to read As-Is DOT after generation
# ---------------------------------------------------------------------------

def get_dot_path(diagram_type: str) -> Path | None:
    """Return the expected DOT file path for *diagram_type*, or None if unknown."""
    entry = _DIAGRAM_MAP.get(diagram_type)
    if not entry:
        return None
    return _REPO_ROOT / entry[3]  # dot_filename stem (no extension)


# ---------------------------------------------------------------------------
# DOT → Mermaid converter
# ---------------------------------------------------------------------------

def dot_to_mermaid(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to Mermaid flowchart syntax.

    Handles:
      - Directed (digraph) and undirected (graph) graphs
      - Node labels (quoted strings)
      - Edge labels
      - Node fill-colour → Mermaid style classes
      - Basic subgraph blocks (flattened)

    Guarantees: Returns syntactically valid Mermaid with escaped special chars.
    """
    lines: list[str] = ["flowchart TD"]

    # Strip comments
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)

    # Determine direction
    if "rankdir=LR" in dot or "rankdir = LR" in dot:
        lines[0] = "flowchart LR"

    node_ids: dict[str, str] = {}   # raw_id → safe_id
    node_labels: dict[str, str] = {}
    node_colors: dict[str, str] = {}
    edge_count = 0

    def _safe_id(raw: str) -> str:
        """Convert a DOT node id to a Mermaid-safe alphanumeric id."""
        s = re.sub(r"[^a-zA-Z0-9_]", "_", raw.strip('"'))
        if s and s[0].isdigit():
            s = "N" + s
        return s or "NODE"

    def _safe_label(lbl: str) -> str:
        """Escape label for Mermaid syntax safety."""
        lbl = lbl.strip('"').strip("'")
        # Remove/replace problematic chars for Mermaid
        lbl = lbl.replace("[", "(").replace("]", ")").replace("{", "(").replace("}", ")")
        lbl = lbl.replace('"', "'")  # replace double quotes with single
        # Truncate very long labels
        if len(lbl) > 80:
            lbl = lbl[:77] + "..."
        return lbl

    is_directed = "digraph" in dot.lower()
    arrow = "-->" if is_directed else "---"

    # ── Parse node statements ──────────────────────────────────────────────
    # Pattern: id [label="...", fillcolor="...", ...]
    node_pattern = re.compile(
        r'^\s*"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE
    )
    for m in node_pattern.finditer(dot):
        raw_id = m.group(1)
        attrs_str = m.group(2)
        safe = _safe_id(raw_id)
        node_ids[raw_id] = safe

        # Extract label
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs_str)
        label = _safe_label(lbl_m.group(1)) if lbl_m else _safe_label(raw_id)
        node_labels[raw_id] = label

        # Extract fillcolor for style
        col_m = re.search(r'fillcolor\s*=\s*"?([^",\]]+)"?', attrs_str)
        if col_m:
            node_colors[raw_id] = col_m.group(1).strip()

    # ── Parse edge statements ──────────────────────────────────────────────
    # Pattern: id1 -> id2 [label="..."]  or  id1 -- id2
    edge_sep = r"->" if is_directed else r"--"
    edge_pattern = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pattern.finditer(dot):
        src_raw = m.group(1).strip()
        dst_raw = m.group(2).strip()
        attrs_str = m.group(3) or ""

        src = node_ids.get(src_raw, _safe_id(src_raw))
        dst = node_ids.get(dst_raw, _safe_id(dst_raw))

        # Ensure nodes are declared
        if src_raw not in node_labels:
            node_labels[src_raw] = _safe_label(src_raw)
            node_ids[src_raw] = src
        if dst_raw not in node_labels:
            node_labels[dst_raw] = _safe_label(dst_raw)
            node_ids[dst_raw] = dst

        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs_str)
        edge_label = _safe_label(lbl_m.group(1)) if lbl_m else ""

        src_label = node_labels.get(src_raw, src)
        dst_label = node_labels.get(dst_raw, dst)

        if edge_label:
            lines.append(f'    {src}["{src_label}"] {arrow}|"{edge_label}"| {dst}["{dst_label}"]')
        else:
            lines.append(f'    {src}["{src_label}"] {arrow} {dst}["{dst_label}"]')
        edge_count += 1

    # ── Emit isolated nodes (no edges) ────────────────────────────────────
    for raw_id, safe in node_ids.items():
        label = node_labels.get(raw_id, raw_id)
        # Only add if not already in an edge line
        node_ref = f'{safe}['
        if not any(node_ref in l for l in lines[1:]):
            lines.append(f'    {safe}["{label}"]')

    # ── Style classes from fillcolors ──────────────────────────────────────
    color_classes: dict[str, list[str]] = {}
    for raw_id, color in node_colors.items():
        safe = node_ids.get(raw_id, _safe_id(raw_id))
        color_classes.setdefault(color, []).append(safe)

    for color, ids in color_classes.items():
        cls_name = "cls_" + re.sub(r"[^a-zA-Z0-9]", "_", color)
        lines.append(f'    classDef {cls_name} fill:{color},stroke:#333,color:#fff')
        lines.append(f'    class {",".join(ids)} {cls_name}')

    if edge_count == 0 and not node_ids:
        # Fallback: no parseable structure
        lines.append('    A["Diagram data not parsed"]')

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# DOT → PlantUML converter
# ---------------------------------------------------------------------------

def dot_to_plantuml(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to PlantUML component/activity syntax."""
    is_directed = "digraph" in dot_text.lower()

    # Strip comments
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)

    puml_type = "component" if diagram_type == "technology" else "activity" if diagram_type == "process" else "usecase"

    lines = ["@startuml", f"' {diagram_type.title()} Architecture Diagram", ""]

    if diagram_type == "people":
        lines += ["left to right direction", ""]

    node_labels: dict[str, str] = {}
    edges: list[tuple[str, str, str]] = []

    def _label(raw: str) -> str:
        return raw.strip('"').strip("'").replace("<", "").replace(">", "")

    # Parse nodes
    node_pat = re.compile(r'"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE)
    for m in node_pat.finditer(dot):
        raw_id = m.group(1)
        attrs = m.group(2)
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        label = _label(lbl_m.group(1)) if lbl_m else _label(raw_id)
        node_labels[raw_id] = label

    # Parse edges
    edge_sep = "->" if is_directed else "--"
    edge_pat = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pat.finditer(dot):
        src = m.group(1).strip()
        dst = m.group(2).strip()
        attrs = m.group(3) or ""
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        elabel = _label(lbl_m.group(1)) if lbl_m else ""
        edges.append((src, dst, elabel))
        if src not in node_labels:
            node_labels[src] = _label(src)
        if dst not in node_labels:
            node_labels[dst] = _label(dst)

    arrow = "-->" if is_directed else "--"

    for raw_id, label in node_labels.items():
        safe_id = re.sub(r"[^a-zA-Z0-9]", "_", raw_id)
        if diagram_type == "technology":
            lines.append(f'component "{label}" as {safe_id}')
        elif diagram_type == "people":
            lines.append(f'actor "{label}" as {safe_id}')
        else:  # process — use rectangle so the full network structure is preserved
            lines.append(f'rectangle "{label}" as {safe_id}')

    lines.append("")

    for src, dst, elabel in edges:
        s = re.sub(r"[^a-zA-Z0-9]", "_", src)
        d = re.sub(r"[^a-zA-Z0-9]", "_", dst)
        lbl_part = f" : {elabel}" if elabel else ""
        lines.append(f"{s} {arrow} {d}{lbl_part}")

    lines.append("")
    lines.append("@enduml")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# DOT → draw.io XML converter
# ---------------------------------------------------------------------------

def dot_to_drawio(dot_text: str, diagram_type: str = "process") -> str:
    """Convert a Graphviz DOT string to draw.io (diagrams.net) XML.

    The returned XML string can be imported directly into draw.io / diagrams.net
    via File → Import from → XML, or saved as a ``.drawio`` file and opened.
    """
    dot = re.sub(r"//[^\n]*", "", dot_text)
    dot = re.sub(r"/\*.*?\*/", "", dot, flags=re.DOTALL)
    is_directed = "digraph" in dot.lower()

    def _lbl(raw: str) -> str:
        return raw.strip('"').strip("'")

    node_labels: dict[str, str] = {}
    node_colors: dict[str, str] = {}
    edges: list[tuple[str, str, str]] = []

    node_pat = re.compile(r'"?([^"\s\[{};]+)"?\s*\[([^\]]*)\]', re.MULTILINE)
    for m in node_pat.finditer(dot):
        raw_id, attrs = m.group(1), m.group(2)
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        node_labels[raw_id] = _lbl(lbl_m.group(1)) if lbl_m else _lbl(raw_id)
        col_m = re.search(r'fillcolor\s*=\s*"?([^",\]]+)"?', attrs)
        if col_m:
            node_colors[raw_id] = col_m.group(1).strip()

    edge_sep = "->" if is_directed else "--"
    edge_pat = re.compile(
        r'"?([^"\s\[{};>-]+)"?\s*' + re.escape(edge_sep) + r'\s*"?([^"\s\[{};>-]+)"?'
        r'(?:\s*\[([^\]]*)\])?',
        re.MULTILINE,
    )
    for m in edge_pat.finditer(dot):
        src, dst = m.group(1).strip(), m.group(2).strip()
        attrs = m.group(3) or ""
        lbl_m = re.search(r'label\s*=\s*"([^"]*)"', attrs)
        edges.append((src, dst, _lbl(lbl_m.group(1)) if lbl_m else ""))
        node_labels.setdefault(src, _lbl(src))
        node_labels.setdefault(dst, _lbl(dst))

    # Style per diagram type
    _STYLE = {
        "process":    "rounded=1;whiteSpace=wrap;html=1;fillColor=#dae8fc;strokeColor=#6c8ebf;",
        "people":     "shape=mxgraph.basic.person;whiteSpace=wrap;html=1;fillColor=#f8cecc;strokeColor=#b85450;",
        "technology": "rounded=1;arcSize=20;whiteSpace=wrap;html=1;fillColor=#d5e8d4;strokeColor=#82b366;",
    }
    base_style = _STYLE.get(diagram_type, _STYLE["process"])
    edge_style = (
        "rounded=0;orthogonalLoop=1;jettySize=auto;exitX=0.5;exitY=1;entryX=0.5;entryY=0;"
        + ("endArrow=block;endFill=1;" if is_directed else "endArrow=none;")
    )

    # Layout: left-right for technology, top-down otherwise
    cols = max(1, int(len(node_labels) ** 0.5) + 1)
    x_gap, y_gap = 160, 100
    x_start, y_start = 80, 80

    cells: list[str] = [
        '<mxCell id="0"/>',
        '<mxCell id="1" parent="0"/>',
    ]
    node_ids: dict[str, str] = {}
    cell_id = 2

    for i, (raw_id, label) in enumerate(node_labels.items()):
        col, row = i % cols, i // cols
        x, y = x_start + col * x_gap, y_start + row * y_gap
        color_part = (
            f'fillColor={node_colors[raw_id]};'
            if raw_id in node_colors else ""
        )
        cid = str(cell_id)
        node_ids[raw_id] = cid
        safe_label = label.replace('"', '&quot;').replace('<', '&lt;').replace('>', '&gt;')
        cells.append(
            f'<mxCell id="{cid}" value="{safe_label}" style="{base_style}{color_part}" '
            f'vertex="1" parent="1"><mxGeometry x="{x}" y="{y}" width="120" height="60" '
            f'as="geometry"/></mxCell>'
        )
        cell_id += 1

    for src, dst, elabel in edges:
        src_id = node_ids.get(src, "1")
        dst_id = node_ids.get(dst, "1")
        safe_elabel = elabel.replace('"', '&quot;')
        cells.append(
            f'<mxCell id="{cell_id}" value="{safe_elabel}" style="{edge_style}" '
            f'edge="1" source="{src_id}" target="{dst_id}" parent="1">'
            f'<mxGeometry relative="1" as="geometry"/></mxCell>'
        )
        cell_id += 1

    cells_xml = "\n        ".join(cells)
    title = f"{diagram_type.title()} Architecture"
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<mxfile version="20.0.0">\n'
        f'  <diagram name="{title}">\n'
        f'    <mxGraphModel dx="1034" dy="546" grid="1" gridSize="10" guides="1" '
        f'tooltips="1" connect="1" arrows="1" fold="1" page="0" pageScale="1" '
        f'pageWidth="1169" pageHeight="827" math="0" shadow="0">\n'
        f'      <root>\n'
        f'        {cells_xml}\n'
        f'      </root>\n'
        f'    </mxGraphModel>\n'
        f'  </diagram>\n'
        f'</mxfile>'
    )


# ---------------------------------------------------------------------------
# Mermaid syntax validation
# ---------------------------------------------------------------------------

def _validate_mermaid(mmd_text: str) -> tuple[bool, str]:
    """Validate Mermaid syntax. Returns (is_valid, error_msg)."""
    lines = [l.strip() for l in mmd_text.split('\n') if l.strip()]
    if not lines:
        return False, "Empty diagram"

    # Check first line is a valid diagram declaration
    first = lines[0].lower()
    if not any(first.startswith(d) for d in ['flowchart', 'graph', 'sequencediagram', 'classDiagram']):
        return False, f"Invalid diagram type: {first}"

    # Basic bracket matching
    open_brackets = mmd_text.count('[') + mmd_text.count('(')
    close_brackets = mmd_text.count(']') + mmd_text.count(')')
    if open_brackets != close_brackets:
        return False, f"Unbalanced brackets: {open_brackets} open, {close_brackets} close"

    return True, ""


# ---------------------------------------------------------------------------
# Main executor
# ---------------------------------------------------------------------------

def run(
    diagram_type: str,
    excel_path: str | None = None,
    output_format: str = "mermaid",   # "mermaid" only; stored formats generated as artifacts
    is_tobe: bool = False,            # AS-IS (False) or TO-BE (True)
) -> str | None:
    """Generate a QODE diagram and return the Mermaid path.

    Always returns Mermaid .mmd format. Artifacts in other formats (DOT, PlantUML, draw.io)
    are generated for archival but not returned.

    Args:
        diagram_type:   "process" | "people" | "technology"
        excel_path:     Path to the QODE questionnaire. Defaults to repo root default.
        output_format:  Ignored; always returns Mermaid.
        is_tobe:        True = TO-BE diagrams, False = AS-IS diagrams (default).

    Returns:
        Absolute path to the Mermaid .mmd file, or None on failure.
    """
    if diagram_type not in _DIAGRAM_MAP:
        logger.error("Unknown diagram_type '%s'", diagram_type)
        return None

    module_name, class_name, method_name, dot_filename = _DIAGRAM_MAP[diagram_type]

    # Smart questionnaire resolution — accepts any .xlsm, not just the hardcoded name
    resolved_excel = find_questionnaire(excel_path)
    if resolved_excel is None:
        logger.error(
            "No QODE questionnaire file found. "
            "Upload a .xlsm/.xlsx file or place one in the repo root."
        )
        return None
    excel_abs = resolved_excel

    original_cwd = os.getcwd()
    os.chdir(str(_REPO_ROOT))

    # Ensure the generator modules always find the file under the canonical name
    expected_path = _REPO_ROOT / "QODE-Questionnaire.xlsm"
    if Path(excel_abs).resolve() != expected_path.resolve() and Path(excel_abs).exists():
        import shutil
        shutil.copy(excel_abs, str(expected_path))

    try:
        repo_str = str(_REPO_ROOT)
        if repo_str not in sys.path:
            sys.path.insert(0, repo_str)

        if module_name in sys.modules:
            del sys.modules[module_name]

        import importlib
        module = importlib.import_module(module_name)
        instance = getattr(module, class_name)()
        getattr(instance, method_name)()

        dot_path = _REPO_ROOT / dot_filename
        if not dot_path.exists():
            logger.warning("DOT file '%s' was not created.", dot_path)
            return None

        dot_text = dot_path.read_text(encoding="utf-8", errors="replace")

        output_dirs = TOBE_DIRS if is_tobe else ASIS_DIRS

        # ── Save DOT to structured dir ────────────────────────────────────
        try:
            dot_out = output_dirs["dot"] / f"{dot_filename}.dot"
            dot_out.write_text(dot_text, encoding="utf-8")
        except Exception as _e:
            logger.warning("Could not write DOT to structured dir: %s", _e)

        # ── Primary path: Mermaid (.mmd) ──────────────────────────────────
        mmd_out = output_dirs["mermaid"] / f"{dot_filename}.mmd"
        try:
            mmd_text = dot_to_mermaid(dot_text, diagram_type)
            is_valid, err_msg = _validate_mermaid(mmd_text)
            if not is_valid:
                logger.error("Mermaid validation failed: %s", err_msg)
                return None
            mmd_out.write_text(mmd_text, encoding="utf-8")
            logger.info("Mermaid diagram saved: %s", mmd_out)

            # Generate artifact formats (not returned to UI)
            # ── PlantUML artifact
            try:
                puml_text = dot_to_plantuml(dot_text, diagram_type)
                puml_out = output_dirs["plantuml"] / f"{dot_filename}.puml"
                puml_out.write_text(puml_text, encoding="utf-8")
                logger.info("PlantUML artifact saved: %s", puml_out)
            except Exception:
                pass

            # ── draw.io artifact
            try:
                drawio_text = dot_to_drawio(dot_text, diagram_type)
                drawio_out = output_dirs["drawio"] / f"{dot_filename}.drawio"
                drawio_out.write_text(drawio_text, encoding="utf-8")
                logger.info("draw.io artifact saved: %s", drawio_out)
            except Exception:
                pass

            return str(mmd_out)
        except Exception as mmd_err:
            logger.error("Mermaid conversion failed: %s", mmd_err)
            return None

    except Exception as exc:
        logger.error("Diagram generation failed for '%s': %s", diagram_type, exc)
        return None

    finally:
        os.chdir(original_cwd)
